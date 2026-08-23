from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import defaultdict
from collections.abc import Iterable
from dataclasses import dataclass
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook

COVERAGE_FIELDS = [
    "deal_id",
    "party_role",
    "party_name",
    "normalized_employer_alias",
    "fiscal_year",
    "period",
    "certified_case_count",
    "new_employment_sum",
]

NORMALIZATION_RULE = (
    "Uppercase EMPLOYER_NAME, replace every run of non-alphanumeric characters with one "
    "space, and trim leading and trailing spaces."
)


@dataclass(frozen=True)
class Alias:
    deal_id: str
    party_role: str
    party_name: str
    normalized_employer_alias: str


def normalize_employer_name(value: object) -> str:
    """Apply the prespecified employer-name normalization, without fuzzy matching."""
    return re.sub(r"[^A-Z0-9]+", " ", str(value or "").upper()).strip()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_aliases(path: Path) -> list[Alias]:
    with path.open(newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        expected = {"deal_id", "party_role", "party_name", "normalized_employer_alias"}
        if not reader.fieldnames or not expected.issubset(reader.fieldnames):
            raise ValueError(f"Alias CSV must contain: {', '.join(sorted(expected))}")
        aliases = [
            Alias(
                deal_id=(row["deal_id"] or "").strip(),
                party_role=(row["party_role"] or "").strip(),
                party_name=(row["party_name"] or "").strip(),
                normalized_employer_alias=normalize_employer_name(row["normalized_employer_alias"]),
            )
            for row in reader
        ]
    if any(not alias.deal_id or not alias.normalized_employer_alias for alias in aliases):
        raise ValueError("Every alias row must have a deal_id and employer alias")
    keys = [(alias.deal_id, alias.normalized_employer_alias) for alias in aliases]
    if len(keys) != len(set(keys)):
        raise ValueError("Alias CSV contains a duplicate deal_id/employer alias")
    return aliases


def _load_windows(review_csv: Path, deal_ids: set[str]) -> dict[str, tuple[int, int]]:
    candidates: dict[str, tuple[int, int]] = {}
    with review_csv.open(newline="", encoding="utf-8-sig") as file:
        reader = csv.DictReader(file)
        required = {"deal_id", "announcement_date", "effective_date"}
        if not reader.fieldnames or not required.issubset(reader.fieldnames):
            raise ValueError(f"Pilot review CSV must contain: {', '.join(sorted(required))}")
        for row in reader:
            deal_id = (row["deal_id"] or "").strip()
            if deal_id not in deal_ids:
                continue
            try:
                window = (
                    date.fromisoformat((row["announcement_date"] or "").strip()).year - 1,
                    date.fromisoformat((row["effective_date"] or "").strip()).year + 1,
                )
            except ValueError as error:
                raise ValueError(
                    f"Deal {deal_id} requires valid announcement/effective dates"
                ) from error
            if deal_id in candidates and candidates[deal_id] != window:
                raise ValueError(f"Deal {deal_id} has conflicting event dates")
            candidates[deal_id] = window
    missing = sorted(deal_ids - candidates.keys())
    if missing:
        raise ValueError(f"Pilot review CSV is missing alias deals: {', '.join(missing)}")
    return candidates


def _decimal(value: object) -> Decimal:
    if value is None or str(value).strip() == "":
        return Decimal(0)
    try:
        return Decimal(str(value).replace(",", "").strip())
    except InvalidOperation as error:
        raise ValueError(f"Invalid NEW_EMPLOYMENT value: {value!r}") from error


def _display_decimal(value: Decimal) -> str:
    if value == value.to_integral():
        return str(int(value))
    return format(value.normalize(), "f")


def _workbook_totals(path: Path, aliases: set[str]) -> dict[str, tuple[int, Decimal]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip().upper() for value in next(rows)]
        except StopIteration as error:
            raise ValueError(f"Workbook has no header row: {path}") from error
        required = {"EMPLOYER_NAME", "CASE_STATUS", "NEW_EMPLOYMENT"}
        if not required.issubset(headers):
            missing = ", ".join(sorted(required - set(headers)))
            raise ValueError(f"Workbook {path} is missing columns: {missing}")
        positions = {name: headers.index(name) for name in required}
        totals: dict[str, tuple[int, Decimal]] = defaultdict(lambda: (0, Decimal(0)))
        for row in rows:
            status = str(row[positions["CASE_STATUS"]] or "").strip().upper()
            if not status.startswith("CERTIFIED"):
                continue
            employer = normalize_employer_name(row[positions["EMPLOYER_NAME"]])
            if employer not in aliases:
                continue
            count, positions_total = totals[employer]
            totals[employer] = (
                count + 1,
                positions_total + _decimal(row[positions["NEW_EMPLOYMENT"]]),
            )
        return dict(totals)
    finally:
        workbook.close()


def _write_csv(path: Path, rows: Iterable[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=COVERAGE_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def audit_h1b_coverage(
    review_csv: Path,
    aliases_csv: Path,
    workbooks: dict[int, Path],
    output_dir: Path,
) -> dict[str, object]:
    """Audit local official LCA workbooks. This function performs no network access."""
    if not workbooks:
        raise ValueError("Provide at least one fiscal-year workbook")
    aliases = _load_aliases(aliases_csv)
    windows = _load_windows(review_csv, {alias.deal_id for alias in aliases})
    needed_years = {year for window in windows.values() for year in window}
    missing_years = sorted(needed_years - workbooks.keys())
    if missing_years:
        raise ValueError(
            "Missing workbooks for required fiscal years: "
            + ", ".join(str(year) for year in missing_years)
        )

    alias_names = {alias.normalized_employer_alias for alias in aliases}
    totals_by_year = {
        year: _workbook_totals(path, alias_names) for year, path in sorted(workbooks.items())
    }
    coverage_rows: list[dict[str, object]] = []
    deal_totals: dict[tuple[str, str], tuple[int, Decimal]] = defaultdict(lambda: (0, Decimal(0)))
    for alias in aliases:
        pre_year, post_year = windows[alias.deal_id]
        for fiscal_year in sorted(workbooks):
            period = (
                "pre" if fiscal_year == pre_year else "post" if fiscal_year == post_year else ""
            )
            count, new_employment = totals_by_year[fiscal_year].get(
                alias.normalized_employer_alias, (0, Decimal(0))
            )
            coverage_rows.append(
                {
                    "deal_id": alias.deal_id,
                    "party_role": alias.party_role,
                    "party_name": alias.party_name,
                    "normalized_employer_alias": alias.normalized_employer_alias,
                    "fiscal_year": fiscal_year,
                    "period": period,
                    "certified_case_count": count,
                    "new_employment_sum": _display_decimal(new_employment),
                }
            )
            if period:
                prior_count, prior_positions = deal_totals[(alias.deal_id, period)]
                deal_totals[(alias.deal_id, period)] = (
                    prior_count + count,
                    prior_positions + new_employment,
                )

    summaries = []
    for deal_id in sorted(windows):
        pre_year, post_year = windows[deal_id]
        pre_count, pre_positions = deal_totals[(deal_id, "pre")]
        post_count, post_positions = deal_totals[(deal_id, "post")]
        summaries.append(
            {
                "deal_id": deal_id,
                "pre_fiscal_year": pre_year,
                "post_fiscal_year": post_year,
                "pre_certified_case_count": pre_count,
                "post_certified_case_count": post_count,
                "pre_new_employment_sum": _display_decimal(pre_positions),
                "post_new_employment_sum": _display_decimal(post_positions),
                "both_period_case_presence": pre_count > 0 and post_count > 0,
                "both_period_positive_new_employment": pre_positions > 0 and post_positions > 0,
            }
        )

    output_dir.mkdir(parents=True, exist_ok=True)
    _write_csv(output_dir / "h1b_pilot_coverage.csv", coverage_rows)
    manifest: dict[str, object] = {
        "artifact": "H-1B LCA pilot coverage audit",
        "offline_only": True,
        "interpretation": (
            "NEW_EMPLOYMENT is an employer-reported application field, not verified hiring. "
            "Certified case presence is not a measure of realized hires, retention, total hiring, "
            "worker welfare, or a causal acquisition effect."
        ),
        "broad_hiring_outcome_decision": "no-go",
        "rules": {
            "employer_name_normalization": NORMALIZATION_RULE,
            "alias_matching": "Exact match after normalization; no fuzzy matching.",
            "case_filter": "CASE_STATUS stripped, uppercased, and begins with CERTIFIED.",
            "case_measure": "Count matching workbook rows.",
            "new_employment_measure": "Sum NEW_EMPLOYMENT; blank values contribute zero.",
            "pre_fiscal_year": "Announcement calendar year minus one.",
            "post_fiscal_year": "Effective/closing calendar year plus one.",
        },
        "inputs": {
            "pilot_review_queue": {
                "path": str(review_csv),
                "sha256": _sha256(review_csv),
            },
            "alias_crosswalk": {"path": str(aliases_csv), "sha256": _sha256(aliases_csv)},
            "official_fy_q4_workbooks": [
                {"fiscal_year": year, "path": str(path), "sha256": _sha256(path)}
                for year, path in sorted(workbooks.items())
            ],
        },
        "deal_summaries": summaries,
        "deal_count": len(summaries),
        "deals_with_both_period_case_presence": sum(
            bool(summary["both_period_case_presence"]) for summary in summaries
        ),
        "deals_with_both_period_positive_new_employment": sum(
            bool(summary["both_period_positive_new_employment"]) for summary in summaries
        ),
    }
    (output_dir / "h1b_pilot_coverage_manifest.json").write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    return manifest
